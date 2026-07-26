#!/usr/bin/env python3
"""
Seragamkan header kolom keputusan di template Disassembly SIAP.

Banyak template Powertrain memakai REUSE | SALVG | REPAIR, padahal kolom
ketiga artinya "ganti part baru". Parser OCMS hanya mengenali REPLACE, jadi
selama header masih REPAIR kolom itu tidak pernah memicu Part Request.

Perubahan:
  kolom 2 -> SALVAGE  (dari SALVG / SALV)
  kolom 3 -> REPLACE  (dari REPAIR)

Jalankan:
  python tools/normalize_decision_headers.py --dry-run
  python tools/normalize_decision_headers.py
"""
from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

ROOT = Path(__file__).resolve().parents[1]
SIAP_DIRS = [
    ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "ENGINE" / "_SIAP_UPLOAD_GSHEET",
    ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "POWERTRAIN" / "_SIAP_UPLOAD_GSHEET",
]


def norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).upper().strip())


def set_cell(ws, row: int, col: int, value) -> bool:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged in list(ws.merged_cells.ranges):
            if cell.coordinate in merged:
                ws.unmerge_cells(str(merged))
                break
        cell = ws.cell(row, col)
    cell.value = value
    return True


def fix_sheet(ws) -> list[str]:
    """Cari baris header REUSE|SALV*|REPAIR dan seragamkan labelnya."""
    changes: list[str] = []
    max_row = min(ws.max_row, 600)
    max_col = min(ws.max_column, 50)

    for r in range(1, max_row + 1):
        reuse_col = salv_col = repair_col = replace_col = None

        for c in range(1, max_col + 1):
            t = norm(ws.cell(r, c).value)
            if not t:
                continue
            if t in ("REUSE", "REUSED"):
                reuse_col = c
            elif t.startswith("SALV"):
                salv_col = c
            elif t == "REPAIR":
                repair_col = c
            elif t == "REPLACE":
                replace_col = c

        if reuse_col is None or salv_col is None:
            continue

        # SALVG / SALV -> SALVAGE
        if norm(ws.cell(r, salv_col).value) != "SALVAGE":
            old = ws.cell(r, salv_col).value
            set_cell(ws, r, salv_col, "SALVAGE")
            changes.append(f"r{r}c{salv_col}: {old!r} -> 'SALVAGE'")

        # REPAIR di sebelah kanan SALVAGE = kolom ganti baru -> REPLACE
        if replace_col is None and repair_col is not None and repair_col > salv_col:
            set_cell(ws, r, repair_col, "REPLACE")
            changes.append(f"r{r}c{repair_col}: 'REPAIR' -> 'REPLACE'")

    return changes


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--i-know-this-destroys-images",
        action="store_true",
        help="lewati pengaman (jangan dipakai pada template SIAP)",
    )
    args = ap.parse_args()

    if not args.dry_run and not args.i_know_this_destroys_images:
        raise SystemExit(
            "DIHENTIKAN: menyimpan workbook lewat openpyxl MENGHAPUS gambar.\n"
            "Template SIAP berisi gambar EMF/WMF yang tidak didukung openpyxl;\n"
            "sekali disimpan, seluruh drawing hilang (file 14 MB pernah jadi 85 KB).\n\n"
            "Pakai versi Excel COM yang tidak merusak format:\n"
            "  pwsh -File tools/normalize_headers_excel.ps1 -DryRun\n"
            "  pwsh -File tools/normalize_headers_excel.ps1\n"
        )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_root = ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "_SIAP_BACKUP" / f"{ts}_headers"

    total_files = total_changes = 0

    for siap_dir in SIAP_DIRS:
        if not siap_dir.is_dir():
            continue

        for path in sorted(siap_dir.rglob("*.xlsx")):
            if path.name.startswith("~$") or "DISASSEMBLY" not in path.name.upper():
                continue

            wb = openpyxl.load_workbook(path)
            file_changes: dict[str, list[str]] = {}

            for sn in wb.sheetnames:
                changes = fix_sheet(wb[sn])
                if changes:
                    file_changes[sn] = changes

            if file_changes:
                total_files += 1
                rel = str(path.relative_to(ROOT)).replace("\\", "/")
                print(f"\n{rel}")
                for sn, changes in file_changes.items():
                    print(f"  [{sn}]")
                    for ch in changes:
                        print(f"    {ch}")
                        total_changes += 1

                if not args.dry_run:
                    dest = backup_root / path.relative_to(ROOT)
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    if not dest.exists():
                        shutil.copy2(path, dest)
                    wb.save(path)

            wb.close()

    print("\n" + "-" * 72)
    print(f"{'DRY RUN — ' if args.dry_run else ''}File diubah: {total_files} | Sel header: {total_changes}")
    if not args.dry_run and total_files:
        print(f"Backup: {backup_root.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
