#!/usr/bin/env python3
"""
Verifikasi kondisi template SIAP setelah format_siap_templates.py.

Untuk tiap file Disassembly/Inspection di _SIAP_UPLOAD_GSHEET, laporkan:
- tab mana saja yang punya tabel keputusan (penting: OCMS hanya baca SATU tab)
- apakah tab itu akan terpilih oleh keyword auto-read OCMS
- header keputusan yang terdeteksi
- berapa baris part dan berapa sel keputusan yang masih non-boolean

Jalankan:
  python tools/verify_siap_state.py
  python tools/verify_siap_state.py --json tools/siap_state.json
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SIAP_DIRS = [
    ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "ENGINE" / "_SIAP_UPLOAD_GSHEET",
    ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "POWERTRAIN" / "_SIAP_UPLOAD_GSHEET",
]

# Keyword yang dipakai OCMS (ChecksheetGsheetService) untuk memilih tab.
DISASSEMBLY_KEYWORDS = ("disassy", "diss", "disassembly", "engine")
INSPECTION_KEYWORDS = ("inspeksi", "inspection", "measurement")


def norm(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).upper().strip())


def is_part_number(value) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if s.upper() in ("NO", "NO."):
        return False
    return bool(re.match(r"^\d+$", s.split(".")[0]))


# Sama persis dengan isDecisionChecked() di ChecksheetGsheetService.php.
CHECKED_VALUES = {
    "x", "✓", "√", "v", "yes", "y", "true", "1", "●", "☑", "✔", "checked", "ya",
}


def would_count_as_checked(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return False
    if isinstance(value, (int, float)):
        return float(value) == 1.0
    return str(value).strip().lower() in CHECKED_VALUES


def scan_sheet(ws, profile: str) -> dict | None:
    """Cari tabel keputusan di satu tab. None kalau tidak ada."""
    max_row = min(ws.max_row, 600)
    max_col = min(ws.max_column, 50)

    headers: list[dict] = []

    for r in range(1, max_row + 1):
        labels: dict[str, int] = {}
        name_col = no_col = None

        for c in range(1, max_col + 1):
            t = norm(ws.cell(r, c).value)
            if not t:
                continue
            if re.match(r"^NO\.?$", t):
                no_col = c
            # Samakan dengan detectDisassemblyHeaderColumns() di
            # app/Services/ChecksheetGsheetService.php
            if re.search(r"PARTS?\s*TO\s*REMOVE|PARTS?\s*NAME|^PARTS?$", t):
                name_col = c
            if profile == "disassembly":
                if t in ("REUSE", "REUSED"):
                    labels["REUSE"] = c
                elif t.startswith("SALV"):
                    labels["SALVAGE"] = c
                elif t == "REPLACE":
                    labels["REPLACE"] = c
                elif t == "REPAIR":
                    labels.setdefault("REPAIR(!)", c)
            else:
                if t in ("U/A", "UA"):
                    labels["U/A"] = c
                elif t in ("U/R", "UR"):
                    labels["U/R"] = c
                elif t in ("R/N", "RN"):
                    labels["R/N"] = c

        # Sub-header keputusan bisa 1-2 baris di bawah header nama (merged DECISION).
        if name_col is not None and not labels:
            for rr in range(r + 1, min(r + 3, max_row + 1)):
                for c in range(1, max_col + 1):
                    t = norm(ws.cell(rr, c).value)
                    if profile == "inspection":
                        if t in ("U/A", "UA"):
                            labels["U/A"] = c
                        elif t in ("U/R", "UR"):
                            labels["U/R"] = c
                        elif t in ("R/N", "RN"):
                            labels["R/N"] = c
                    else:
                        if t in ("REUSE", "REUSED"):
                            labels["REUSE"] = c
                        elif t.startswith("SALV"):
                            labels["SALVAGE"] = c
                        elif t == "REPLACE":
                            labels["REPLACE"] = c
                if labels:
                    break

        if name_col is not None and labels:
            headers.append({"row": r, "name_col": name_col, "no_col": no_col, "labels": labels})

    if not headers:
        return None

    # Hitung baris part + sel keputusan non-boolean, dimulai dari header terakhir
    # (layout nyata: satu tabel per tab untuk inspection, multi-section untuk disassembly).
    part_rows = 0
    stray_checked = 0   # teks di kolom keputusan yang AKAN dianggap centang
    stray_text = 0      # teks lain — parser mengabaikannya
    blank_cells = 0

    for idx, h in enumerate(headers):
        decision_cols = [c for k, c in h["labels"].items() if not k.endswith("(!)")]
        if not decision_cols:
            continue
        name_col = h["name_col"]
        no_col = h["no_col"] or max(1, name_col - 1)

        # Berhenti di header berikutnya, jangan sampai baris yang sama
        # dihitung ulang oleh setiap section.
        end_row = headers[idx + 1]["row"] - 1 if idx + 1 < len(headers) else max_row

        for r in range(h["row"] + 1, end_row + 1):
            no_val = ws.cell(r, no_col).value
            name_val = ws.cell(r, name_col).value
            if not is_part_number(no_val) or not name_val or not str(name_val).strip():
                continue
            part_rows += 1
            for c in decision_cols:
                val = ws.cell(r, c).value
                if val is None:
                    blank_cells += 1
                elif not isinstance(val, bool):
                    if would_count_as_checked(val):
                        stray_checked += 1
                    else:
                        stray_text += 1

    return {
        "headers": len(headers),
        "labels": sorted({k for h in headers for k in h["labels"]}),
        "part_rows": part_rows,
        "stray_checked": stray_checked,
        "stray_text": stray_text,
        "blank_cells": blank_cells,
    }


def tabs_read_by_ocms(sheetnames: list[str], keywords: tuple[str, ...]) -> list[str]:
    """OCMS membaca SEMUA tab yang namanya cocok keyword; kalau tidak ada yang
    cocok, webapp jatuh ke tab pertama."""
    matched = [sn for sn in sheetnames if any(k in sn.lower() for k in keywords)]
    if matched:
        return matched
    return sheetnames[:1]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", dest="json_out")
    args = ap.parse_args()

    results = []

    for siap_dir in SIAP_DIRS:
        if not siap_dir.is_dir():
            print(f"SKIP missing: {siap_dir}")
            continue

        for path in sorted(siap_dir.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            upper = path.name.upper()
            if "DISASSEMBLY" in upper:
                profile = "disassembly"
                keywords = DISASSEMBLY_KEYWORDS
            elif "INSPECTION" in upper:
                profile = "inspection"
                keywords = INSPECTION_KEYWORDS
            else:
                continue

            rel = str(path.relative_to(ROOT)).replace("\\", "/")
            try:
                wb = openpyxl.load_workbook(path, data_only=False)
            except Exception as e:  # noqa: BLE001
                results.append({"file": rel, "profile": profile, "error": str(e)})
                continue

            selected = tabs_read_by_ocms(wb.sheetnames, keywords)
            tabs = {}
            for sn in wb.sheetnames:
                info = scan_sheet(wb[sn], profile)
                if info:
                    tabs[sn] = info
            wb.close()

            results.append(
                {
                    "file": rel,
                    "profile": profile,
                    "sheetnames": wb.sheetnames,
                    "ocms_reads_tabs": selected,
                    "tabs_with_decisions": tabs,
                }
            )

    # Laporan
    print("=" * 78)
    print("STATE TEMPLATE SIAP — kesiapan otomasi FR/PR")
    print("=" * 78)

    issues: list[str] = []   # menghambat otomasi FR/PR
    notes: list[str] = []    # informasi, tidak menghambat

    for r in results:
        if r.get("error"):
            print(f"\nERROR {r['file']}: {r['error']}")
            issues.append(f"{r['file']}: gagal dibaca")
            continue

        tabs = r["tabs_with_decisions"]
        selected = r["ocms_reads_tabs"]
        is_engine = r["file"].split("/")[1] == "ENGINE" if "/" in r["file"] else False

        print(f"\n{r['file']}  [{r['profile']}]")
        print(f"  tab dibaca OCMS : {', '.join(selected)}")

        if not tabs:
            print("  !! tidak ada tabel keputusan terdeteksi")
            # Sumber keputusan Powertrain adalah sheet Inspection, jadi
            # Disassembly tanpa kolom keputusan tidak menghambat FR.
            if r["profile"] == "disassembly" and not is_engine:
                notes.append(f"{r['file']}: tanpa kolom keputusan (Powertrain pakai Inspection)")
            else:
                issues.append(f"{r['file']}: tanpa tabel keputusan")
            continue

        for sn, info in tabs.items():
            mark = "<== dibaca" if sn in selected else "    TIDAK dibaca"
            print(
                f"  [{sn}] {mark} headers={info['headers']} labels={','.join(info['labels'])} "
                f"parts={info['part_rows']} centang-liar={info['stray_checked']} "
                f"teks={info['stray_text']} kosong={info['blank_cells']}"
            )
            if info["stray_checked"]:
                issues.append(
                    f"{r['file']} [{sn}]: {info['stray_checked']} sel keputusan berisi nilai "
                    f"yang akan dibaca sebagai CENTANG"
                )
            if info["stray_text"]:
                notes.append(
                    f"{r['file']} [{sn}]: {info['stray_text']} sel teks di kolom keputusan "
                    f"(diabaikan parser)"
                )
            if any(k.endswith("(!)") for k in info["labels"]):
                issues.append(f"{r['file']} [{sn}]: header REPAIR non-standar")

        unread = [sn for sn in tabs if sn not in selected]
        if unread:
            lost = sum(tabs[sn]["part_rows"] for sn in unread)
            issues.append(
                f"{r['file']}: {len(unread)} tab berisi keputusan TIDAK dibaca OCMS "
                f"({', '.join(unread)}) — {lost} baris part hilang"
            )

    print("\n" + "=" * 78)
    print(f"MASALAH ({len(issues)})  — menghambat otomasi FR/PR")
    print("=" * 78)
    for i, msg in enumerate(issues, 1):
        print(f"{i:3d}. {msg}")

    print("\n" + "=" * 78)
    print(f"CATATAN ({len(notes)})  — tidak menghambat")
    print("=" * 78)
    for i, msg in enumerate(notes, 1):
        print(f"{i:3d}. {msg}")

    if args.json_out:
        Path(args.json_out).write_text(
            json.dumps(
                {"results": results, "issues": issues, "notes": notes},
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        print(f"\nJSON: {args.json_out}")


if __name__ == "__main__":
    main()
