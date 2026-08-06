"""Ekstrak isi checksheet .docx (tabel + heading) menjadi .xlsx BARU.

Dipakai untuk checksheet assembly yang hanya tersedia dalam format Word
(ekspor HTML Word macet karena dokumen penuh drawing). Gambar ilustrasi
tidak ikut — hanya struktur tabel checklist.

Catatan: openpyxl di sini MENULIS FILE BARU dari nol, bukan mengedit
template .xlsx ber-gambar (yang dilarang karena merusak EMF/WMF).

Pakai:
    python tools/docx_tables_to_xlsx.py "file1.docx" "file2.docx"
"""
import sys
from pathlib import Path

from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


def iter_body(doc):
    """Paragraf & tabel sesuai urutan dokumen."""
    body = doc.element.body
    for child in body.iterchildren():
        if child.tag.endswith('}p'):
            yield Paragraph(child, doc)
        elif child.tag.endswith('}tbl'):
            yield Table(child, doc)


def convert(src: Path) -> Path:
    doc = Document(str(src))
    wb = Workbook()
    ws = wb.active
    ws.title = 'CHECKSHEET'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_idx = 1
    n_tables = 0

    for block in iter_body(doc):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            cell = ws.cell(row=row_idx, column=1, value=text)
            cell.font = Font(bold=True)
            row_idx += 1
        else:
            n_tables += 1
            for r in block.rows:
                col_idx = 1
                seen = set()
                for c in r.cells:
                    # Sel merge muncul berulang; tulis sekali saja per objek sel
                    if id(c._tc) in seen:
                        col_idx += 1
                        continue
                    seen.add(id(c._tc))
                    text = '\n'.join(p.text for p in c.paragraphs).strip()
                    cell = ws.cell(row=row_idx, column=col_idx, value=text or None)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    col_idx += 1
                row_idx += 1
            row_idx += 1  # baris kosong antar tabel

    # Lebar kolom kasar supaya terbaca
    for col in ws.columns:
        letter = col[0].column_letter
        longest = max((len(str(c.value).split('\n')[0]) for c in col if c.value), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 8), 60)

    dst = src.with_suffix('.xlsx')
    wb.save(str(dst))
    print(f'OK  {dst.name}: {n_tables} tabel, {row_idx - 1} baris')
    return dst


if __name__ == '__main__':
    for arg in sys.argv[1:]:
        convert(Path(arg))
