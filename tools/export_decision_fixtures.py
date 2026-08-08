#!/usr/bin/env python3
"""
Ekspor nilai sel tab keputusan dari template SIAP ke JSON fixture,
supaya parser PHP bisa diuji dengan data nyata tanpa memanggil Google Sheets.

Bentuk output sama dengan balasan webapp action=read versi baru:
  { "sheets": [ {"name": "...", "values": [[...], ...]} ], "matched": true }

Jalankan:
  python tools/export_decision_fixtures.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "tests" / "Fixtures" / "gsheet"

# (nama fixture, path relatif, keyword tab seperti yang dikirim OCMS)
TARGETS = [
    (
        "inspection_cv_pc1250.json",
        "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY/POWERTRAIN/_SIAP_UPLOAD_GSHEET/"
        "Control Valve/PC1250-8/INSPECTION Control Valve PC1250-8.xlsx",
        ("inspeksi", "inspection", "measurement"),
    ),
    (
        "inspection_cv_d375.json",
        "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY/POWERTRAIN/_SIAP_UPLOAD_GSHEET/"
        "Control Valve/D375-6/INSPECTION Control Valve D375-6.xlsx",
        ("inspeksi", "inspection", "measurement"),
    ),
    (
        "disassembly_subassy_engine_d375.json",
        "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY/ENGINE/_SIAP_UPLOAD_GSHEET/"
        "SUBASSY DISASSEMBLY ENGINE SAA6D170E-5 (D375-6 PC1250-8).xlsx",
        ("disassy", "diss", "disassembly", "engine"),
    ),
]

MAX_ROWS = 400
MAX_COLS = 40


def cell_value(v):
    """Samakan dengan yang dikirim Apps Script: primitif JSON."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for fixture_name, rel_path, keywords in TARGETS:
        path = ROOT / rel_path
        if not path.exists():
            print(f"SKIP missing: {rel_path}")
            continue

        wb = openpyxl.load_workbook(path, data_only=False)
        sheets = []
        for sn in wb.sheetnames:
            low = sn.lower()
            if not any(k in low for k in keywords):
                continue
            ws = wb[sn]
            rows = []
            for r in range(1, min(ws.max_row, MAX_ROWS) + 1):
                rows.append([cell_value(ws.cell(r, c).value) for c in range(1, min(ws.max_column, MAX_COLS) + 1)])
            sheets.append({"name": sn, "values": rows})
        wb.close()

        payload = {"ok": True, "sheets": sheets, "matched": len(sheets) > 0}
        if sheets:
            payload["sheet"] = sheets[0]["name"]
            payload["values"] = sheets[0]["values"]

        out = OUT_DIR / fixture_name
        out.write_text(json.dumps(payload), encoding="utf-8")
        print(f"{fixture_name}: {len(sheets)} tab -> {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
