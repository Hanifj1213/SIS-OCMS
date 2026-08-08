#!/usr/bin/env python3
"""
Format template SIAP (_SIAP_UPLOAD_GSHEET) untuk otomasi FR/PR:

- Disassembly: header REUSE | SALVAGE | REPLACE + sel decision = FALSE (siap checkbox GSheet)
- Inspection:  header U/A | U/R | R/N + sel decision = FALSE
- Perbaiki outlier D375 (SALVG) dan HD785 (tanpa kolom decision)

Jalankan:
  python tools/format_siap_templates.py --dry-run
  python tools/format_siap_templates.py
"""
from __future__ import annotations

import argparse
import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SIAP_DIRS = [
    ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "ENGINE" / "_SIAP_UPLOAD_GSHEET",
    ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "POWERTRAIN" / "_SIAP_UPLOAD_GSHEET",
]

SKIP_NAME_PARTS = ("MEASUREMENT", "RECEIVING", "DELIVERY", "TEST")  # not ASSEMBLY — substring of DISASSEMBLY

BOOL_DV = DataValidation(
    type="list",
    formula1='"FALSE,TRUE"',
    allow_blank=True,
    showDropDown=True,
)
BOOL_DV.error = "Pilih FALSE (kosong) atau TRUE (centang)"
BOOL_DV.errorTitle = "Keputusan Part"


@dataclass
class DecisionColumns:
    no_col: int
    name_col: int
    decision_cols: dict[str, int]  # label -> 0-based col index
    header_row: int  # 1-based


@dataclass
class SheetReport:
    sheet: str
    profile: str
    headers_found: int = 0
    part_rows: int = 0
    cells_formatted: int = 0
    fixes: list[str] = field(default_factory=list)


@dataclass
class FileReport:
    path: str
    sheets: list[SheetReport] = field(default_factory=list)
    skipped: str | None = None
    error: str | None = None


def safe_set_cell(ws, row: int, col: int, value) -> None:
    """Tulis sel; unmerge dulu jika merged (header template sering di-merge)."""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged in list(ws.merged_cells.ranges):
            if cell.coordinate in merged:
                ws.unmerge_cells(str(merged))
                break
        cell = ws.cell(row, col)
    cell.value = value


def norm(cell) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell).upper().strip())


def is_part_number(value) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if s.upper() in ("NO", "NO."):
        return False
    try:
        int(s.split(".")[0])
        return True
    except ValueError:
        return bool(re.match(r"^\d+$", s))


def sheet_matches(name: str, keywords: tuple[str, ...]) -> bool:
    low = name.lower()
    return any(k in low for k in keywords)


def detect_disassembly_header(row_values: list, row_idx: int) -> DecisionColumns | None:
    no_col = name_col = None
    reuse_col = salvage_col = replace_col = None

    for c, cell in enumerate(row_values):
        t = norm(cell)
        if not t:
            continue
        if re.match(r"^NO\.?$", t):
            no_col = c
        if re.search(r"PARTS TO REMOVE|PARTS?\s*NAME|^PARTS$", t):
            name_col = c
        if t in ("REUSE", "REUSED"):
            reuse_col = c
        if t.startswith("SALV") or t in ("SALVAGE", "SALVG"):
            salvage_col = c
        if t == "REPLACE":
            replace_col = c
        if t == "REPAIR" and salvage_col is None and replace_col is None:
            # D375 first REPAIR col — handled by special fix
            pass

    if name_col is None:
        return None

    if no_col is None:
        no_col = max(0, name_col - 1)

    # Standard triple
    if reuse_col is not None and salvage_col is not None and replace_col is not None:
        return DecisionColumns(
            no_col=no_col,
            name_col=name_col,
            decision_cols={"REUSE": reuse_col, "SALVAGE": salvage_col, "REPLACE": replace_col},
            header_row=row_idx,
        )

    # Salvage-only variant (partial) — still usable
    if salvage_col is not None:
        cols = {}
        if reuse_col is not None:
            cols["REUSE"] = reuse_col
        cols["SALVAGE"] = salvage_col
        if replace_col is not None:
            cols["REPLACE"] = replace_col
        return DecisionColumns(
            no_col=no_col,
            name_col=name_col,
            decision_cols=cols,
            header_row=row_idx,
        )

    return None


def detect_inspection_header(rows: list[list], start: int) -> DecisionColumns | None:
    """Need header row + sub-header row for U/A U/R R/N."""
    if start >= len(rows):
        return None

    header = rows[start]
    no_col = name_col = None
    for c, cell in enumerate(header):
        t = norm(cell)
        if re.match(r"^NO\.?$", t):
            no_col = c
        if re.search(r"PARTS?\s*NAME", t):
            name_col = c

    if name_col is None:
        return None
    if no_col is None:
        no_col = max(0, name_col - 1)

    decision_cols: dict[str, int] = {}
    for r in range(start, min(start + 3, len(rows))):
        for c, cell in enumerate(rows[r]):
            t = norm(cell)
            if t in ("U/A", "UA"):
                decision_cols["U/A"] = c
            elif t in ("U/R", "UR") or "U/R" in t:
                decision_cols["U/R"] = c
            elif t in ("R/N", "RN") or "R/N" in t:
                decision_cols["R/N"] = c

    if not decision_cols:
        return None

    return DecisionColumns(
        no_col=no_col,
        name_col=name_col,
        decision_cols=decision_cols,
        header_row=start + 1,
    )


def apply_special_disassembly_fixes(ws, rel_path: str, report: SheetReport) -> None:
    p = rel_path.replace("\\", "/").upper()

    # D375 CV: REPAIR | SALVG | REPAIR  →  REUSE | SALVAGE | REPLACE
    if "CONTROL VALVE/D375-6/DISASSEMBLY" in p:
        for r in range(1, min(40, ws.max_row + 1)):
            v18 = norm(ws.cell(r, 18).value)
            v19 = norm(ws.cell(r, 19).value)
            if v18 == "REPAIR" and v19.startswith("SALV"):
                safe_set_cell(ws, r, 18, "REUSE")
                safe_set_cell(ws, r, 19, "SALVAGE")
                safe_set_cell(ws, r, 20, "REPLACE")
                report.fixes.append(f"Baris {r}: REPAIR/SALVG -> REUSE/SALVAGE/REPLACE")

    if "CONTROL VALVE/HD785-7/DISASSEMBLY" in p:
        for r in range(1, min(40, ws.max_row + 1)):
            v3 = norm(ws.cell(r, 3).value)
            v4 = norm(ws.cell(r, 4).value)
            if v3 in ("NO.", "NO") and v4 == "PARTS":
                safe_set_cell(ws, r, 18, "REUSE")
                safe_set_cell(ws, r, 19, "SALVAGE")
                safe_set_cell(ws, r, 20, "REPLACE")
                safe_set_cell(ws, r, 21, "SIGN")
                # Sub-header tanda tangan geser ke kolom 21-23
                sub = r + 1
                if sub <= ws.max_row:
                    if ws.cell(sub, 18).value:
                        safe_set_cell(ws, sub, 21, ws.cell(sub, 18).value)
                        safe_set_cell(ws, sub, 18, None)
                    if ws.cell(sub, 20).value:
                        safe_set_cell(ws, sub, 22, ws.cell(sub, 20).value)
                        safe_set_cell(ws, sub, 20, None)
                    if ws.cell(sub, 21).value and norm(ws.cell(sub, 21).value) == "SPV":
                        safe_set_cell(ws, sub, 23, ws.cell(sub, 21).value)
                        safe_set_cell(ws, sub, 21, None)
                report.fixes.append(f"Baris {r}: tambah REUSE/SALVAGE/REPLACE (HD785)")


def format_decision_cells(ws, cols: DecisionColumns, report: SheetReport, dry_run: bool) -> None:
    decision_indices = list(cols.decision_cols.values())
    if not decision_indices:
        return

    start_row = cols.header_row + 1
    in_section = True

    for r in range(start_row, ws.max_row + 1):
        row_vals = [ws.cell(r, c + 1).value for c in range(max(decision_indices + [cols.name_col, cols.no_col]) + 2)]

        # New header section?
        detected = detect_disassembly_header(row_vals, r) if "REUSE" in cols.decision_cols or "SALVAGE" in cols.decision_cols else None
        if detected and r > cols.header_row + 1:
            cols = detected
            decision_indices = list(cols.decision_cols.values())
            report.headers_found += 1
            continue

        no_val = ws.cell(r, cols.no_col + 1).value
        name_val = ws.cell(r, cols.name_col + 1).value

        if not is_part_number(no_val) or not name_val or not str(name_val).strip():
            continue

        report.part_rows += 1

        if dry_run:
            report.cells_formatted += len(decision_indices)
            continue

        for c in decision_indices:
            cell = ws.cell(r, c + 1)
            if isinstance(cell, MergedCell):
                continue
            # Hanya reset teks bebas; biarkan FALSE/TRUE/checkbox yang sudah benar
            val = cell.value
            if val is None or (isinstance(val, str) and val.strip() and val.strip().upper() not in ("TRUE", "FALSE")):
                cell.value = False
                report.cells_formatted += 1
            elif val is True:
                cell.value = True
            elif val is False:
                cell.value = False

        # Data validation TRUE/FALSE (dropdown — di GSheet bisa diganti Insert Checkbox)
        min_c = min(decision_indices) + 1
        max_c = max(decision_indices) + 1
        range_ref = f"{openpyxl.utils.get_column_letter(min_c)}{r}:{openpyxl.utils.get_column_letter(max_c)}{r}"
        dv = copy(BOOL_DV)
        dv.add(range_ref)
        ws.add_data_validation(dv)

    # Catatan di sel header pertama decision
    if not dry_run and cols.decision_cols:
        first_label = next(iter(cols.decision_cols))
        first_col = cols.decision_cols[first_label] + 1
        hdr = ws.cell(cols.header_row, first_col)
        if hdr.comment is None:
            hdr.comment = Comment(
                "Set FALSE lalu di Google Sheets: Insert → Checkbox. Centang SALVAGE/U/R untuk FR.",
                "OCMS",
            )


def format_inspection_sheet(ws, report: SheetReport, dry_run: bool) -> None:
    max_scan = min(ws.max_row, 500)
    rows = []
    for r in range(1, max_scan + 1):
        rows.append([ws.cell(r, c).value for c in range(1, min(ws.max_column, 40) + 1)])

    for r in range(len(rows)):
        cols = detect_inspection_header(rows, r)
        if cols is None:
            continue

        report.headers_found += 1
        report.profile = "inspection"

        decision_indices = list(cols.decision_cols.values())
        for dr in range(cols.header_row + 1, ws.max_row + 1):
            no_val = ws.cell(dr, cols.no_col + 1).value
            name_val = ws.cell(dr, cols.name_col + 1).value

            if not is_part_number(no_val) or not name_val or not str(name_val).strip():
                continue

            # Stop at next table header
            if norm(no_val) == "NO" or norm(name_val).startswith("PARTS"):
                break

            report.part_rows += 1
            if dry_run:
                report.cells_formatted += len(decision_indices)
                continue

            for c in decision_indices:
                cell = ws.cell(dr, c + 1)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None or (isinstance(val, str) and val.strip() and val.strip().upper() not in ("TRUE", "FALSE")):
                    cell.value = False
                    report.cells_formatted += 1

            min_c = min(decision_indices) + 1
            max_c = max(decision_indices) + 1
            range_ref = f"{openpyxl.utils.get_column_letter(min_c)}{dr}:{openpyxl.utils.get_column_letter(max_c)}{dr}"
            dv = copy(BOOL_DV)
            dv.add(range_ref)
            ws.add_data_validation(dv)


def format_disassembly_sheet(ws, rel_path: str, report: SheetReport, dry_run: bool) -> None:
    report.profile = "disassembly"
    apply_special_disassembly_fixes(ws, rel_path, report)

    max_scan = min(ws.max_row, 800)
    headers: list[DecisionColumns] = []

    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 50) + 1)]
        detected = detect_disassembly_header(row_vals, r)
        if detected:
            headers.append(detected)

    if not headers:
        report.fixes.append("Header REUSE/SALVAGE/REPLACE tidak ditemukan")
        return

    report.headers_found = len(headers)

    for cols in headers:
        format_decision_cells(ws, cols, report, dry_run)


def should_process_file(path: Path) -> tuple[bool, str]:
    name = path.name.upper()
    if "DISASSEMBLY" in name or "INSPECTION" in name:
        return True, ""
    if "MEASUREMENT" in name or "RECEIVING" in name or "DELIVERY" in name or "TEST" in name:
        return False, "skip stage (measurement/receiving/delivery/test)"
    if "ASSEMBLY" in name:
        return False, "skip assembly"
    return False, "bukan disassembly/inspection"


def process_workbook(path: Path, rel: str, dry_run: bool) -> FileReport:
    fr = FileReport(path=rel)
    ok, reason = should_process_file(path)
    if not ok:
        fr.skipped = reason
        return fr

    name = path.name.upper()
    is_inspection = "INSPECTION" in name

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        fr.error = str(e)
        return fr

    for sn in wb.sheetnames:
        if is_inspection and not sheet_matches(sn, ("inspeksi", "inspection")):
            # PC1250 multi-tab: proses semua tab inspeksi
            if not sheet_matches(sn, ("inspeksi", "inspection", "tm", "pm")):
                continue
        if not is_inspection and "DISASSEMBLY" in name:
            if not sheet_matches(sn, ("disassy", "diss", "disassembly", "engine")) and len(wb.sheetnames) > 1:
                continue

        ws = wb[sn]
        sr = SheetReport(sheet=sn, profile="inspection" if is_inspection else "disassembly")

        if is_inspection:
            format_inspection_sheet(ws, sr, dry_run)
        else:
            format_disassembly_sheet(ws, rel, sr, dry_run)

        if sr.headers_found or sr.fixes or sr.part_rows:
            fr.sheets.append(sr)

    if not dry_run and fr.sheets and not fr.error:
        wb.save(path)

    wb.close()
    return fr


def backup_file(path: Path, backup_root: Path) -> None:
    rel = path.relative_to(ROOT)
    dest = backup_root / rel
    dest.parent.mkdir(parents=True, exist_ok=True)
    if not dest.exists():
        shutil.copy2(path, dest)


def main() -> None:
    parser = argparse.ArgumentParser(description="Format SIAP templates for FR/PR automation")
    parser.add_argument("--dry-run", action="store_true", help="Hanya laporan, tidak menulis file")
    parser.add_argument("--no-backup", action="store_true", help="Jangan buat backup")
    parser.add_argument(
        "--i-know-this-destroys-images",
        action="store_true",
        help="lewati pengaman (jangan dipakai pada template SIAP)",
    )
    args = parser.parse_args()

    if not args.dry_run and not args.i_know_this_destroys_images:
        raise SystemExit(
            "DIHENTIKAN: menyimpan workbook lewat openpyxl MENGHAPUS gambar.\n"
            "Menjalankan skrip ini pada 2026-07-26 merusak 30 file: gambar EMF/WMF\n"
            "dibuang dan drawing ikut hilang (DISASSEMBLY D375-6: 14,6 MB -> 85 KB).\n\n"
            "Pengisian sel FALSE + data validation per baris juga TIDAK diperlukan:\n"
            "sel kosong sudah dibaca sebagai 'tidak dicentang' oleh parser OCMS, dan\n"
            "Google Sheets mengubah sel kosong jadi checkbox saat Insert > Checkbox.\n\n"
            "Yang benar-benar perlu hanya penyeragaman label header:\n"
            "  pwsh -File tools/normalize_headers_excel.ps1\n\n"
            "Cek kondisi template: python tools/verify_siap_state.py\n"
            "Cek kerusakan format: python tools/inspect_xlsx_damage.py\n"
        )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_root = ROOT / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY" / "_SIAP_BACKUP" / ts

    reports: list[FileReport] = []

    for siap_dir in SIAP_DIRS:
        if not siap_dir.is_dir():
            print(f"SKIP missing: {siap_dir}")
            continue

        for path in sorted(siap_dir.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue

            rel = str(path.relative_to(ROOT)).replace("\\", "/")

            if not args.dry_run and not args.no_backup:
                backup_file(path, backup_root)

            fr = process_workbook(path, rel, args.dry_run)
            reports.append(fr)

    # Summary
    print("\n" + "=" * 72)
    print(f"FORMAT SIAP TEMPLATES {'(DRY RUN)' if args.dry_run else ''}")
    if not args.dry_run and not args.no_backup:
        print(f"Backup: {backup_root.relative_to(ROOT)}")
    print("=" * 72)

    processed = skipped = errors = 0
    total_parts = total_cells = 0

    for fr in reports:
        if fr.error:
            errors += 1
            print(f"\nERROR {fr.path}\n  {fr.error}")
            continue
        if fr.skipped:
            skipped += 1
            continue

        processed += 1
        print(f"\n{fr.path}")
        for sr in fr.sheets:
            print(f"  [{sr.sheet}] {sr.profile}: headers={sr.headers_found}, parts={sr.part_rows}, cells={sr.cells_formatted}")
            for fix in sr.fixes:
                print(f"    fix: {fix}")
            total_parts += sr.part_rows
            total_cells += sr.cells_formatted

    print("\n" + "-" * 72)
    print(f"Processed: {processed} | Skipped: {skipped} | Errors: {errors}")
    print(f"Part rows touched: {total_parts} | Decision cells formatted: {total_cells}")
    print("\nSetelah upload ke Google Sheets: Insert > Checkbox di kolom decision (TRUE/FALSE -> checkbox).")


if __name__ == "__main__":
    main()
