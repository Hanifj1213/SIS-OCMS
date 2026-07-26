"""
Build Control Valve receiving quiz items (grouped + callout numbers)
from COMPLETED workbooks → database/data/control_valve_receiving_items.json
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
COMPLETED = (
    ROOT
    / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY"
    / "POWERTRAIN"
    / "CHECKSHEET CONTROL VALVE POWERTRAIN ALL UNIT PT 2026"
)
OUT = ROOT / "database" / "data" / "control_valve_receiving_items.json"

FILES = {
    "D155-6": "cs cv pm d155-6(COMPLETED).xlsx",
    "D375-6": "cs cv pm d375-6(COMPLETED).xlsx",
    "HD785-7": "cs cv tf hd785-7(COMPLETED).xlsx",
    "GD825A-2": "cs cv tm gd825-2(COMPLETED).xlsx",
    "WA800-3": "cs cv tm wa800-3(COMPLETED).xlsx",
}

SECTION_KEYS = (
    "MAIN RELIEF",
    "ECMV",
    "STEERING",
    "TORQ",
    "TRANSMISSION",
    "LUBRICATING",
    "STATOR",
    "LOCK",
    "FILTER",
    "REGULATOR",
    "CONTROL VALVE",
    "INCHING",
)

SKIP_NAMES = {
    "NO",
    "ITEM NAME",
    "ITEM",
    "SKETCH DRAWING OR PHOTOS",
    "CHECK",
    "REMARK",
    "RECEIVING BY",
    "PART NAME",
    "PART NUMBER",
    "DOCUMENT NO",
    "PUBLISHED DATE",
    "REVISION NO",
    "PAGE",
    "RO",
    "SN",
    "UNIT MODEL",
    "UNIT CODE",
    "START DATE / TIME",
    "FINISH DATE / TIME",
}


def find_receiving(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "receiving":
            return wb[name]
    return wb[wb.sheetnames[0]]


def is_section(text: str) -> bool:
    up = text.upper().strip()
    if len(up) < 8:
        return False
    if up in SKIP_NAMES:
        return False
    if "UNTUK ECMV" in up:  # note rows
        return False
    return any(k in up for k in SECTION_KEYS) and not re.search(r"\(\d+\s*PCS", up, re.I)


def parse_sheet(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_receiving(wb)
    section = "Visual Inspection"
    items: list[dict] = []
    seen: set[tuple[str, int, str]] = set()

    for r in range(1, (ws.max_row or 0) + 1):
        cells: list[tuple[int, object]] = []
        for c in range(1, min(24, (ws.max_column or 0) + 1)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if isinstance(v, str):
                v = re.sub(r"\s+", " ", v.strip())
                if not v:
                    continue
            cells.append((c, v))
        if not cells:
            continue

        # section header: a long title without a numeric NO in typical item cols
        for _, v in cells:
            if isinstance(v, str) and is_section(v):
                # avoid treating item names as sections
                if re.search(r"\(\d+\s*PCS", v, re.I):
                    continue
                if re.match(r"^\d+", v):
                    continue
                section = v.upper().strip()
                break

        no = None
        for _, v in cells:
            if isinstance(v, (int, float)) and 1 <= int(v) <= 40:
                no = int(v)
                break
            if isinstance(v, str) and re.fullmatch(r"\d{1,2}", v.strip()):
                no = int(v.strip())
                break

        if no is None:
            continue

        name = None
        for _, v in cells:
            if not isinstance(v, str):
                continue
            up = v.upper().strip()
            if up in SKIP_NAMES or "SKETCH" in up:
                continue
            if is_section(v) and not re.search(r"\(\d+\s*PCS|/ WRAP|/ WRAPING|PAINTING|PACKING", up):
                continue
            if len(v) < 4:
                continue
            if name is None or len(v) > len(name):
                name = v

        if not name:
            continue

        key = (section, no, name.upper())
        if key in seen:
            continue
        seen.add(key)

        items.append(
            {
                "group": section.title() if section != "Visual Inspection" else section,
                "number": no,
                "label": name,
            }
        )

    wb.close()
    return items


def main() -> None:
    payload = {}
    for egi, fn in FILES.items():
        path = COMPLETED / fn
        items = parse_sheet(path)
        # assign sequential ids
        for i, it in enumerate(items, start=1):
            it["id"] = f"CVL-{i:03d}"
        payload[egi] = items
        print(f"{egi}: {len(items)} items")
        for it in items[:5]:
            print(f"  #{it['number']:02d} [{it['group']}] {it['label']}")
        if len(items) > 5:
            print("  ...")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nWrote {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
