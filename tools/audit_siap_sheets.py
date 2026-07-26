#!/usr/bin/env python3
"""Audit ENGINE & POWERTRAIN _SIAP_UPLOAD_GSHEET workbooks for standardization."""
import os
import re
import json
from collections import defaultdict

import openpyxl

ROOT = os.path.join(
    os.path.dirname(__file__),
    "..",
    "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY",
)

FOLDERS = {
    "ENGINE": os.path.join(ROOT, "ENGINE", "_SIAP_UPLOAD_GSHEET"),
    "POWERTRAIN": os.path.join(ROOT, "POWERTRAIN", "_SIAP_UPLOAD_GSHEET"),
}

STAGE_PATTERNS = [
    (r"RECEIVING", "Receiving"),
    (r"DISASSEMBLY|DISASSY|DISS", "Disassembly"),
    (r"INSPECTION|INSPEKSI|MEASUREMENT|MEASURING", "Inspection/Measurement"),
    (r"ASSEMBLY|ASSY", "Assembly"),
    (r"DELIVERY|DELIVER", "Delivery"),
    (r"TEST", "Test"),
    (r"SUBASSY", "Subassy"),
]

DECISION_KEYWORDS = {
    "reuse_salvage_replace": ["REUSE", "SALVAGE", "REPLACE"],
    "u_decision": ["U/A", "U/R", "R/N", "U/A | U/R"],
    "good_bad": ["NO GOOD", "GOOD", "OK", "NOT OK"],
    "new_reuse": ["REPLACE NEW", "REUSED", "NEW"],
    "repair_limit": ["REPAIR LIMIT"],
}


def classify_stage(filename: str) -> str:
    upper = filename.upper()
    for pattern, label in STAGE_PATTERNS:
        if re.search(pattern, upper):
            return label
    return "Other"


def norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).upper().strip())


def scan_workbook(path: str) -> dict:
    rel = path
    for base in FOLDERS.values():
        if path.startswith(base):
            rel = os.path.relpath(path, base)
            break

    info = {
        "file": rel.replace("\\", "/"),
        "stage": classify_stage(os.path.basename(path)),
        "sheets": [],
        "decision_profile": None,
        "part_list_profile": None,
        "issues": [],
    }

    size_mb = os.path.getsize(path) / (1024 * 1024)
    info["size_mb"] = round(size_mb, 2)
    if size_mb > 8:
        info["issues"].append(f"File besar ({size_mb:.1f} MB) — scan header saja")
        return info

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        info["issues"].append(f"Cannot open: {e}")
        return info

    info["sheet_count"] = len(wb.sheetnames)
    info["sheet_names"] = wb.sheetnames[:12]

    profiles = defaultdict(int)

    try:
        for sn in wb.sheetnames[:15]:
            ws = wb[sn]
            max_row = min(getattr(ws, "max_row", 0) or 0, 80)
            max_col = min(getattr(ws, "max_column", 0) or 0, 35)

            found = {k: [] for k in DECISION_KEYWORDS}
            header_hits = defaultdict(list)

            try:
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
                    for c, cell in enumerate(row, start=1):
                        if cell is None:
                            continue
                        t = norm(cell)
                        if not t:
                            continue

                        for group, kws in DECISION_KEYWORDS.items():
                            for kw in kws:
                                if kw in t or t == kw.replace(" ", ""):
                                    found[group].append(t[:40])

                        if re.match(r"^NO\.?$", t):
                            header_hits["no"].append(1)
                        if re.search(r"PARTS?\s*NAME|PARTS TO REMOVE", t):
                            header_hits["part_name"].append(1)
                        if "DECISION" in t:
                            header_hits["decision"].append(1)
            except Exception as e:
                info["issues"].append(f"Sheet {sn}: {e}")
                continue

            sheet_info = {
                "name": sn,
                "size": f"{getattr(ws, 'max_row', '?')}x{getattr(ws, 'max_column', '?')}",
            }

            sheet_profile = "none"
            if found["u_decision"]:
                sheet_profile = "U/A|U/R|R/N"
            elif found["reuse_salvage_replace"]:
                sheet_profile = "REUSE|SALVAGE|REPLACE"
            elif found["new_reuse"]:
                sheet_profile = "NEW|REUSED|REPLACE NEW"
            elif found["good_bad"]:
                sheet_profile = "GOOD|NO GOOD"
            elif found["repair_limit"]:
                sheet_profile = "measurement_only"

            sheet_info["decision_profile"] = sheet_profile
            sheet_info["has_part_list"] = bool(header_hits["no"] and header_hits["part_name"])
            profiles[sheet_profile] += 1

            if sheet_profile != "none" or sheet_info["has_part_list"]:
                info["sheets"].append(sheet_info)
    finally:
        wb.close()

    if profiles:
        info["decision_profile"] = max(profiles, key=profiles.get)
        if len(profiles) > 1:
            info["decision_profiles_mixed"] = dict(profiles)

    multi = [s for s in info["sheets"] if s.get("decision_profile") != "none"]
    part_sheets = [s for s in info["sheets"] if s.get("has_part_list")]
    info["part_list_profile"] = "yes" if part_sheets else "no"

    if info["stage"] == "Inspection/Measurement" and info.get("decision_profile") == "measurement_only":
        info["issues"].append("Inspection sheet tanpa kolom keputusan part (hanya ukuran)")

    if info["stage"] == "Disassembly" and info.get("decision_profile") != "REUSE|SALVAGE|REPLACE":
        info["issues"].append("Disassembly tanpa REUSE/SALVAGE/REPLACE konsisten")

    if info["sheet_count"] > 1 and info["stage"] == "Inspection/Measurement":
        mixed = info.get("decision_profiles_mixed")
        if mixed and len(mixed) > 1:
            info["issues"].append(f"Multi-tab measurement dengan format campur: {mixed}")

    return info


def main():
    all_results = {"ENGINE": [], "POWERTRAIN": []}

    for group, folder in FOLDERS.items():
        if not os.path.isdir(folder):
            print(f"Missing: {folder}")
            continue
        for root, _, files in os.walk(folder):
            for f in sorted(files):
                if not f.lower().endswith((".xlsx", ".xls")):
                    continue
                if f.startswith("~$"):
                    continue
                path = os.path.join(root, f)
                all_results[group].append(scan_workbook(path))

    # Summaries
    for group, items in all_results.items():
        print("\n" + "=" * 72)
        print(f" {group} — {len(items)} files")
        print("=" * 72)

        by_stage = defaultdict(list)
        for it in items:
            by_stage[it["stage"]].append(it)

        for stage, files in sorted(by_stage.items()):
            print(f"\n## {stage} ({len(files)} files)")
            profiles = defaultdict(int)
            for f in files:
                p = f.get("decision_profile") or "unknown"
                profiles[p] += 1
            print("  Decision profiles:", dict(profiles))

            for f in files:
                flags = []
                if f.get("decision_profiles_mixed"):
                    flags.append("MIXED-TABS")
                if f.get("issues"):
                    flags.append("ISSUES")
                flag = " [" + ", ".join(flags) + "]" if flags else ""
                sc = f.get("sheet_count", "?")
                dp = f.get("decision_profile") or "-"
                print(f"  - {f['file']}{flag}")
                print(f"      sheets={sc}, decision={dp}, part_list={f.get('part_list_profile')}")
                if f.get("issues"):
                    for iss in f["issues"]:
                        print(f"      ! {iss}")
                if f.get("decision_profiles_mixed"):
                    print(f"      tabs: {f['decision_profiles_mixed']}")

    out = os.path.join(os.path.dirname(__file__), "siap_sheets_audit.json")
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(all_results, fh, indent=2, ensure_ascii=False)
    print(f"\nFull JSON: {out}")


if __name__ == "__main__":
    main()
