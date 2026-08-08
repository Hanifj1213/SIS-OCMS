#!/usr/bin/env python3
"""Scan all SIAP xlsx for RH/LH/FRONT/CENTRE/REAR + decision headers (read-only)."""
import re
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent / "CHECKSHEET FOR PROCESS DEVELOPMEN ROBBY"
DIRS = [
    ROOT / "ENGINE" / "_SIAP_UPLOAD_GSHEET",
    ROOT / "POWERTRAIN" / "_SIAP_UPLOAD_GSHEET",
]

SIDE = {"RH", "LH"}
POS = {"FRONT", "CENTRE", "CENTER", "REAR"}
DECISION_RSR = {"REUSE", "SALVAGE", "REPLACE"}
DECISION_UR = {"U/A", "U/R", "R/N", "UA", "UR", "RN"}


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).upper().strip())


def scan_file(path: Path) -> dict:
    rel = path.name
    for d in DIRS:
        try:
            rel = str(path.relative_to(d)).replace("\\", "/")
            break
        except ValueError:
            pass

    out = {
        "file": rel,
        "size_mb": round(path.stat().st_size / 1024 / 1024, 2),
        "sheets_with_rhlh": [],
        "decision_type": None,
    }

    if out["size_mb"] > 15:
        out["skip"] = "too large"
        return out

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        out["error"] = str(e)
        return out

    for ws in wb.worksheets:
        has_side = has_pos = False
        decision = set()
        sample_rows = []

        for i, row in enumerate(ws.iter_rows(max_row=min(ws.max_row or 0, 120), max_col=min(ws.max_column or 0, 30), values_only=True)):
            texts = [norm(c) for c in row if norm(c)]
            if not texts:
                continue
            row_set = set(texts)
            if row_set & SIDE:
                has_side = True
            if row_set & POS:
                has_pos = True
            if row_set & DECISION_RSR:
                decision.add("RSR")
            if row_set & DECISION_UR:
                decision.add("UR")

            if (row_set & SIDE) or (row_set & POS):
                if len(sample_rows) < 3:
                    sample_rows.append({"r": i + 1, "cells": texts[:8]})

        if has_side or has_pos:
            out["sheets_with_rhlh"].append({
                "name": ws.title,
                "has_rh_lh": has_side,
                "has_front_centre_rear": has_pos,
                "sample": sample_rows,
            })

    if out["sheets_with_rhlh"]:
        out["decision_type"] = "+".join(sorted(decision)) if decision else None

    wb.close()
    return out


def main():
    files = []
    for d in DIRS:
        if not d.is_dir():
            continue
        files.extend(sorted(d.rglob("*.xlsx")))

    with_rhlh = []
    without = []

    for p in files:
        if p.name.startswith("~$"):
            continue
        info = scan_file(p)
        if info.get("sheets_with_rhlh"):
            with_rhlh.append(info)
        else:
            without.append(info["file"])

    print(f"Total xlsx: {len(files)}")
    print(f"Dengan RH/LH atau FRONT/CENTRE/REAR: {len(with_rhlh)}")
    print(f"Tanpa: {len(without)}\n")

    for info in with_rhlh:
        print(f"FILE: {info['file']}  decision={info.get('decision_type')}")
        for sh in info["sheets_with_rhlh"]:
            flags = []
            if sh["has_rh_lh"]:
                flags.append("RH/LH")
            if sh["has_front_centre_rear"]:
                flags.append("F/C/R")
            print(f"  tab: {sh['name']} [{', '.join(flags)}]")
            for s in sh.get("sample", []):
                print(f"    row {s['r']}: {s['cells']}")
        print()


if __name__ == "__main__":
    main()
